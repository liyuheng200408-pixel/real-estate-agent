#!/usr/bin/env python3
"""每月15日 房源库报表脚本 - 读取悠悠房源库.db，生成Excel发飞书"""

import sqlite3
import os
from pathlib import Path
from datetime import datetime
import sys

# 动态路径：优先读取环境变量，其次用 ~/.hermes/
HERMES_BASE = Path(os.environ.get("HERMES_BASE_DIR", os.path.expanduser("~/.hermes")))
DB_PATH = HERMES_BASE / "data" / "悠悠房源库.db"
OUTPUT_DIR = HERMES_BASE / "cron" / "output"


def get_db_path():
    # 优先使用配置路径，其次用环境变量，最后用默认路径
    return DB_PATH


def query_houses():
    try:
        conn = sqlite3.connect(get_db_path())
        cursor = conn.cursor()
        cursor.execute("""
            SELECT 序号, 标题, 小区, 价格, 单价, 面积, 户型, 楼层, 朝向,
                   装修, 建成年份, 区域, 地址, 经纪人, 评分, 挂牌时间,
                   爬取时间, 状态, 链接
            FROM houses
            ORDER BY 序号
        """)
        rows = cursor.fetchall()
        conn.close()
        return rows
    except Exception as e:
        print(f"数据库读取失败: {e}")
        return []


def generate_excel():
    """生成Excel报表，返回文件路径"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("未安装openpyxl，尝试安装...")
        os.system("pip install openpyxl -q")
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    rows = query_houses()
    if not rows:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "房源库报表"

    # 表头
    headers = ["序号", "标题", "小区", "价格(万)", "单价(元/㎡)", "面积(㎡)",
               "户型", "楼层", "朝向", "装修", "建成年份", "区域", "地址",
               "经纪人", "评分", "挂牌时间", "爬取时间", "状态", "链接"]

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    # 交替行颜色
    even_fill = PatternFill(start_color="D6E3F8", end_color="D6E3F8", fill_type="solid")

    # 写入数据
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if row_idx % 2 == 0:
                cell.fill = even_fill

    # 自动列宽
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 30)

    # 生成文件名
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    today = datetime.now()
    filename = f"儋州房源库_{today.strftime('%Y%m')}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    wb.save(filepath)
    return filepath


if __name__ == "__main__":
    filepath = generate_excel()
    if filepath:
        print(f"Excel已生成: {filepath}")
    else:
        print("生成失败")